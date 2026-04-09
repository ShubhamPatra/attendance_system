"""
Flask routes blueprint -- all web endpoints.
"""

import base64
import collections
import io
import os
from urllib.parse import urlparse
import threading
import tempfile
import time
import uuid
from datetime import datetime

import cv2
import numpy as np
from flask import (
    Blueprint,
    Response,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from redis import Redis

# Magic bytes for image MIME validation
_IMAGE_SIGNATURES = {
    b"\xff\xd8\xff": "jpeg",
    b"\x89PNG": "png",
}


def _is_within_directory(path: str, root: str) -> bool:
  """Return True when *path* is inside *root* after normalization."""
  try:
    return os.path.commonpath([os.path.abspath(path), os.path.abspath(root)]) == os.path.abspath(root)
  except ValueError:
    return False


def _resolve_upload_reference(value: str) -> str | None:
  """Resolve a client-supplied upload reference to a safe absolute path."""
  raw = value.strip()
  if not raw:
    return None

  upload_root = os.path.abspath(config.UPLOAD_DIR)
  if os.path.isabs(raw):
    candidate = os.path.abspath(raw)
  else:
    # Accept basename-style tokens returned by capture endpoint.
    candidate = os.path.abspath(os.path.join(upload_root, os.path.basename(raw)))

  if not _is_within_directory(candidate, upload_root):
    return None
  return candidate


def _validate_image_mime(file_storage) -> str | None:
    """Read magic bytes and return the detected type, or None if invalid."""
    header = file_storage.read(8)
    file_storage.seek(0)
    for sig, fmt in _IMAGE_SIGNATURES.items():
        if header.startswith(sig):
            return fmt
    return None


def _validate_date_param(value: str, name: str) -> tuple[datetime | None, str | None]:
    """Parse a YYYY-MM-DD string.  Returns ``(date, None)`` on success
    or ``(None, error_message)`` on failure."""
    try:
        return datetime.strptime(value, "%Y-%m-%d"), None
    except ValueError:
        return None, f"'{name}' must be a valid date in YYYY-MM-DD format."

import config
import database
from face_engine import encoding_cache, generate_encoding
from performance import tracker
from utils import (
    allowed_file,
    check_image_quality,
    sanitize_string,
    setup_logging,
    today_str,
)

logger = setup_logging()

_rate_limit_lock = threading.Lock()
_rate_limit_buckets: dict[str, collections.deque[float]] = {}


def _check_model_artifacts() -> dict[str, bool]:
  anti_spoof_ok = False
  if os.path.isdir(config.ANTI_SPOOF_MODEL_DIR):
    anti_spoof_ok = any(
      name.lower().endswith(".pth")
      for name in os.listdir(config.ANTI_SPOOF_MODEL_DIR)
    )
  return {
    "yunet_model": os.path.isfile(config.YUNET_MODEL_PATH),
    "anti_spoof_models": anti_spoof_ok,
  }


def _check_mongo_ready() -> bool:
  try:
    database.get_client().admin.command("ping")
    return True
  except Exception:
    return False


def _check_redis_ready() -> bool:
  try:
    parsed = urlparse(config.CELERY_BROKER_URL)
    client = Redis(
      host=parsed.hostname or "localhost",
      port=parsed.port or 6379,
      db=int((parsed.path or "/0").strip("/") or "0"),
      password=parsed.password,
      socket_connect_timeout=2,
      socket_timeout=2,
    )
    client.ping()
    return True
  except Exception:
    return False


def _rate_limit_key(endpoint: str) -> str:
  client = request.headers.get("X-Forwarded-For", request.remote_addr or "unknown")
  client = client.split(",", 1)[0].strip()
  return f"{endpoint}:{client}"


def _check_rate_limit(endpoint: str) -> bool:
  """Return True if the current request is within endpoint rate limits."""
  now = time.monotonic()
  window = max(config.API_RATE_LIMIT_WINDOW_SEC, 1)
  limit = max(config.API_RATE_LIMIT_MAX_REQUESTS, 1)
  key = _rate_limit_key(endpoint)

  with _rate_limit_lock:
    bucket = _rate_limit_buckets.setdefault(key, collections.deque())
    cutoff = now - window
    while bucket and bucket[0] < cutoff:
      bucket.popleft()
    if len(bucket) >= limit:
      return False
    bucket.append(now)
    return True

bp = Blueprint("main", __name__)


# ── Health Endpoints ─────────────────────────────────────────────────────

@bp.route("/health")
def health():
  """Simple liveness endpoint for process-level checks."""
  return jsonify({"status": "ok", "service": "autoattendance"})


@bp.route("/ready")
@bp.route("/healthz")
def ready():
  """Readiness endpoint for container orchestration and probes."""
  checks = {
    "mongo": _check_mongo_ready(),
    "redis": _check_redis_ready(),
    **_check_model_artifacts(),
  }

  # Redis is optional for pure recognition flows, but critical for async jobs.
  critical_ok = checks["mongo"] and checks["yunet_model"] and checks["anti_spoof_models"]
  status_code = 200 if critical_ok else 503
  payload = {
    "status": "ready" if critical_ok else "degraded",
    "checks": checks,
  }
  return jsonify(payload), status_code


# ── Home ──────────────────────────────────────────────────────────────────

@bp.route("/")
def index():
    return render_template("index.html")


# ── Student Registration ─────────────────────────────────────────────────

@bp.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "GET":
        return render_template("register.html")

    # --- Validate form fields ---
    name = sanitize_string(request.form.get("name", ""))
    semester_raw = request.form.get("semester", "")
    reg_no = sanitize_string(request.form.get("registration_number", ""))
    section = sanitize_string(request.form.get("section", ""))

    errors: list[str] = []
    if not name:
        errors.append("Name is required.")
    if not reg_no:
        errors.append("Registration number is required.")
    if not section:
        errors.append("Section is required.")
    try:
        semester = int(semester_raw)
        if semester < 1 or semester > 12:
            errors.append("Semester must be between 1 and 12.")
    except (ValueError, TypeError):
        errors.append("Semester must be a valid integer.")
        semester = None

    # --- Check if webcam mode ---
    webcam_mode = request.form.get("webcam_mode", "0") == "1"

    if webcam_mode:
        # Webcam-based multi-angle enrollment: look for saved frame paths
        webcam_paths: list[str] = []
        for i in range(5):
            ref = request.form.get(f"webcam_frame_{i}", "").strip()
            if not ref:
                continue
            path = _resolve_upload_reference(ref)
            if path is None or not os.path.isfile(path):
                errors.append(f"Webcam frame {i}: invalid file path.")
                continue
            webcam_paths.append(path)
        if not webcam_paths:
            errors.append("No webcam frames provided or files not found.")
    else:
        # --- Validate images (multiple) ---
        files = request.files.getlist("images")
        if not files or all(f.filename == "" for f in files):
            errors.append("Please upload at least one face image.")
        else:
            if len(files) > config.MAX_REGISTRATION_IMAGES:
                errors.append(
                    f"Maximum {config.MAX_REGISTRATION_IMAGES} images allowed."
                )
            for i, f in enumerate(files, 1):
                if f.filename == "":
                    continue
                if not allowed_file(f.filename):
                    errors.append(f"Image {i}: only PNG, JPG, JPEG are allowed.")
                    continue
                # MIME type validation via magic bytes
                detected = _validate_image_mime(f)
                if detected is None:
                    errors.append(
                        f"Image {i}: file content does not match a valid "
                        f"image format (PNG or JPEG)."
                    )
                    continue
                f.seek(0, os.SEEK_END)
                size = f.tell()
                f.seek(0)
                if size > config.UPLOAD_MAX_SIZE:
                    errors.append(f"Image {i}: must be smaller than 5 MB.")

    if errors:
        for e in errors:
            flash(e, "danger")
        return render_template("register.html"), 400

    # --- Process each image: quality check + encoding ---
    encodings: list[np.ndarray] = []

    if webcam_mode:
        for i, path in enumerate(webcam_paths, 1):
            try:
                img_bgr = cv2.imread(path)
                if img_bgr is None:
                    flash(f"Webcam frame {i}: could not read image file.", "danger")
                    continue
                ok, reason = check_image_quality(img_bgr)
                if not ok:
                    flash(f"Webcam frame {i}: {reason}", "danger")
                    continue
                enc = generate_encoding(path)
                encodings.append(enc)
            except ValueError as exc:
                flash(f"Webcam frame {i}: {exc}", "danger")
                continue
    else:
        for i, f in enumerate(files, 1):
            if f.filename == "":
                continue
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(
                    suffix=".jpg", delete=False
                ) as tmp:
                    f.save(tmp)
                    tmp_path = tmp.name

                # Load image for quality checks
                img_bgr = cv2.imread(tmp_path)
                if img_bgr is None:
                    flash(f"Image {i}: could not read image file.", "danger")
                    continue

                # Quality checks (blur + brightness)
                ok, reason = check_image_quality(img_bgr)
                if not ok:
                    flash(f"Image {i}: {reason}", "danger")
                    continue

                # Generate encoding (also checks exactly 1 face)
                enc = generate_encoding(tmp_path)
                encodings.append(enc)
            except ValueError as exc:
                flash(f"Image {i}: {exc}", "danger")
                continue
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    os.unlink(tmp_path)

    if not encodings:
        flash(
            "No valid face encodings could be generated. "
            "Please upload clear, well-lit images with exactly one face.",
            "danger",
        )
        return render_template("register.html"), 400

    # --- Save to database ---
    try:
        database.insert_student(name, semester, reg_no, section, encodings)
    except ValueError as exc:
        flash(str(exc), "danger")
        return render_template("register.html"), 409

    encoding_cache.refresh()
    flash(
        f"Student '{name}' registered successfully with "
        f"{len(encodings)} face encoding(s)!",
        "success",
    )
    return redirect(url_for("main.register"))


# ── Multi-angle Webcam Capture ────────────────────────────────────────────

@bp.route("/api/register/capture", methods=["POST"])
def api_register_capture():
    """Capture a single webcam frame for multi-angle enrollment.
    ---
    tags:
      - Registration
    consumes:
      - application/json
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            frame:
              type: string
              description: Base64-encoded JPEG image data
          required:
            - frame
    responses:
      200:
        description: Frame saved successfully
        schema:
          type: object
          properties:
            path:
              type: string
              description: Filesystem path where the frame was saved
      400:
        description: Missing or invalid frame data
    """
    if not _check_rate_limit("register_capture"):
      return jsonify({"error": "Rate limit exceeded."}), 429

    data = request.get_json(silent=True)
    if not data or "frame" not in data:
        return jsonify({"error": "Missing 'frame' in request body."}), 400

    frame_b64 = data["frame"]
    # Strip optional data URI prefix
    if "," in frame_b64:
        frame_b64 = frame_b64.split(",", 1)[1]

    try:
        raw_bytes = base64.b64decode(frame_b64)
    except Exception:
        return jsonify({"error": "Invalid base64 data."}), 400

    if not raw_bytes.startswith(b"\xff\xd8\xff"):
      return jsonify({"error": "Invalid image data: only JPEG is accepted."}), 400

    arr = np.frombuffer(raw_bytes, dtype=np.uint8)
    if cv2.imdecode(arr, cv2.IMREAD_COLOR) is None:
      return jsonify({"error": "Invalid image data: JPEG decode failed."}), 400

    os.makedirs(config.UPLOAD_DIR, exist_ok=True)
    filename = f"webcam_{uuid.uuid4().hex}.jpg"
    filepath = os.path.join(config.UPLOAD_DIR, filename)

    with open(filepath, "wb") as fh:
        fh.write(raw_bytes)

    return jsonify({"path": filename})


# ── Live Attendance (video feed page) ────────────────────────────────────

@bp.route("/attendance")
def attendance():
    return render_template("attendance.html")


@bp.route("/video_feed")
def video_feed():
    """MJPEG stream of annotated camera frames.
    ---
    tags:
      - Video
    responses:
      200:
        description: MJPEG video stream
        content:
          multipart/x-mixed-replace:
            schema:
              type: string
              format: binary
    """
    from camera import get_camera

    cam = get_camera()

    min_interval = 1.0 / config.MJPEG_TARGET_FPS

    def generate():
        last_yield = 0.0
        while True:
            jpeg = cam.get_latest_jpeg()
            if jpeg is None:
                time.sleep(0.03)
                continue
            # Throttle to target FPS
            now = time.monotonic()
            elapsed = now - last_yield
            if elapsed < min_interval:
                time.sleep(min_interval - elapsed)
            last_yield = time.monotonic()
            yield (
                b"--frame\r\n"
                b"Content-Type: image/jpeg\r\n\r\n" + jpeg + b"\r\n"
            )

    return Response(
        generate(), mimetype="multipart/x-mixed-replace; boundary=frame"
    )


# ── Dashboard ─────────────────────────────────────────────────────────────

@bp.route("/dashboard")
def dashboard():
    total_students = database.student_count()
    today_count = database.today_attendance_count()
    pct = (
        round(today_count / total_students * 100, 1)
        if total_students
        else 0.0
    )
    recent = database.get_attendance()[:20]
    at_risk = database.get_at_risk_students()
    return render_template(
        "dashboard.html",
        total_students=total_students,
        today_count=today_count,
        attendance_pct=pct,
        recent=recent,
        at_risk=at_risk,
    )


# ── Reports ───────────────────────────────────────────────────────────────

@bp.route("/report")
def report():
    date = request.args.get("date", today_str())
    records = database.get_attendance(date)
    return render_template(
        "report.html",
        records=records,
        selected_date=date,
    )


@bp.route("/report/csv")
def report_csv():
    """Export attendance as CSV.
    ---
    tags:
      - Reports
    parameters:
      - name: date
        in: query
        type: string
        required: false
        description: Single date filter (YYYY-MM-DD, default today)
      - name: reg_no
        in: query
        type: string
        required: false
        description: Filter by student registration number
      - name: start_date
        in: query
        type: string
        required: false
        description: Start of date range (YYYY-MM-DD)
      - name: end_date
        in: query
        type: string
        required: false
        description: End of date range (YYYY-MM-DD)
      - name: full
        in: query
        type: string
        required: false
        description: Set to '1' to export entire history
    responses:
      200:
        description: CSV file download
      400:
        description: Invalid date parameters
    """
    reg_no = sanitize_string(request.args.get("reg_no", "").strip())
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    full = request.args.get("full", "").strip()

    if full == "1":
        df = database.get_attendance_csv_full()
        filename = "attendance_full_history.csv"
    elif reg_no:
      if database.get_student_by_reg_no(reg_no) is None:
        return jsonify({"error": "Student not found."}), 404
      df = database.get_attendance_csv_by_student(reg_no)
      filename = f"attendance_{reg_no}.csv"
    elif start_date and end_date:
        sd, err1 = _validate_date_param(start_date, "start_date")
        ed, err2 = _validate_date_param(end_date, "end_date")
        if err1 or err2:
            msg = err1 or err2
            return jsonify({"error": msg}), 400
        if sd > ed:
            return jsonify({"error": "start_date must be <= end_date."}), 400
        df = database.get_attendance_csv_by_date_range(start_date, end_date)
        filename = f"attendance_{start_date}_to_{end_date}.csv"
    else:
        date = request.args.get("date", today_str()).strip()
        _, err = _validate_date_param(date, "date")
        if err:
            return jsonify({"error": err}), 400
        df = database.get_attendance_csv(date)
        filename = f"attendance_{date}.csv"

    buf = io.BytesIO()
    df.to_csv(buf, index=False)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


# ── Excel Export ──────────────────────────────────────────────────────────

@bp.route("/report/xlsx")
def report_xlsx():
    """Export attendance as a formatted Excel (.xlsx) file.
    ---
    tags:
      - Reports
    parameters:
      - name: date
        in: query
        type: string
        required: false
        description: Single date filter (YYYY-MM-DD, default today)
      - name: reg_no
        in: query
        type: string
        required: false
        description: Filter by student registration number
      - name: start_date
        in: query
        type: string
        required: false
        description: Start of date range (YYYY-MM-DD)
      - name: end_date
        in: query
        type: string
        required: false
        description: End of date range (YYYY-MM-DD)
      - name: full
        in: query
        type: string
        required: false
        description: Set to '1' to export entire history
    responses:
      200:
        description: XLSX file download
      400:
        description: Invalid date parameters
    """
    reg_no = sanitize_string(request.args.get("reg_no", "").strip())
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    full = request.args.get("full", "").strip()

    if full == "1":
        df = database.get_attendance_csv_full()
        filename = "attendance_full_history.xlsx"
    elif reg_no:
        df = database.get_attendance_csv_by_student(reg_no)
        filename = f"attendance_{reg_no}.xlsx"
    elif start_date and end_date:
        sd, err1 = _validate_date_param(start_date, "start_date")
        ed, err2 = _validate_date_param(end_date, "end_date")
        if err1 or err2:
            msg = err1 or err2
            return jsonify({"error": msg}), 400
        if sd > ed:
            return jsonify({"error": "start_date must be <= end_date."}), 400
        df = database.get_attendance_csv_by_date_range(start_date, end_date)
        filename = f"attendance_{start_date}_to_{end_date}.xlsx"
    else:
        date = request.args.get("date", today_str()).strip()
        _, err = _validate_date_param(date, "date")
        if err:
            return jsonify({"error": err}), 400
        df = database.get_attendance_csv(date)
        filename = f"attendance_{date}.xlsx"

    # Build formatted workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    columns = list(df.columns)

    # -- Header row styling --
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Freeze the top row
    ws.freeze_panes = "A2"

    # -- Data rows with alternating shading --
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # -- Auto-fit column widths (approximate) --
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── API endpoints ─────────────────────────────────────────────────────────

@bp.route("/api/metrics")
def api_metrics():
    """Return current system performance metrics.
    ---
    tags:
      - Metrics
    responses:
      200:
        description: Performance metrics object
    """
    return jsonify(tracker.metrics())


@bp.route("/api/events")
def api_events():
    """Polling endpoint -- returns recent attendance events as JSON.
    ---
    tags:
      - Events
    responses:
      200:
        description: List of recent attendance events
        schema:
          type: array
          items:
            type: object
    """
    from camera import get_camera
    cam = get_camera()
    return jsonify(cam.pop_events())


# ── Recognition Logs ──────────────────────────────────────────────────────

@bp.route("/logs")
def logs():
    return render_template("logs.html")


@bp.route("/api/logs")
def api_logs():
    """Return recent recognition log entries (non-destructive read).
    ---
    tags:
      - Logs
    responses:
      200:
        description: List of recent log entries
        schema:
          type: array
          items:
            type: object
    """
    from camera import get_camera
    cam = get_camera()
    return jsonify(cam.get_log_buffer())


# ── System Metrics Page ───────────────────────────────────────────────────

@bp.route("/metrics")
def metrics():
    m = tracker.metrics()
    m["total_students"] = database.student_count()
    m["today_count"] = database.today_attendance_count()
    return render_template("metrics.html", metrics=m)


# ── Attendance Activity Visualization ─────────────────────────────────────

@bp.route("/attendance_activity")
def attendance_activity():
    return render_template("attendance_activity.html")


@bp.route("/api/attendance_activity")
def api_attendance_activity():
    """Return attendance counts grouped by hour for a given date.
    ---
    tags:
      - Attendance
    parameters:
      - name: date
        in: query
        type: string
        required: false
        description: Date in YYYY-MM-DD format (default today)
    responses:
      200:
        description: Hourly attendance data
        schema:
          type: object
          properties:
            date:
              type: string
            hours:
              type: array
              items:
                type: object
                properties:
                  hour:
                    type: integer
                  count:
                    type: integer
    """
    date = request.args.get("date", today_str())
    data = database.get_attendance_by_hour(date)
    return jsonify({"date": date, "hours": data})


# ── Registration numbers list (for report filters) ───────────────────────

@bp.route("/api/registration_numbers")
def api_registration_numbers():
    """Return a sorted list of all student registration numbers.
    ---
    tags:
      - Students
    responses:
      200:
        description: List of registration number strings
        schema:
          type: array
          items:
            type: string
    """
    return jsonify(database.get_all_registration_numbers())


# ── Attendance Heatmap ────────────────────────────────────────────────────

@bp.route("/heatmap")
def heatmap():
    return render_template("heatmap.html")


@bp.route("/api/heatmap")
def api_heatmap():
    """Return attendance heatmap data for the last 90 days.
    ---
    tags:
      - Attendance
    responses:
      200:
        description: List of daily attendance counts
        schema:
          type: array
          items:
            type: object
            properties:
              date:
                type: string
              count:
                type: integer
              total_students:
                type: integer
    """
    data = database.get_attendance_heatmap_data()
    return jsonify(data)


# ── Bulk Manual Attendance ────────────────────────────────────────────────

@bp.route("/api/attendance/bulk", methods=["POST"])
def api_attendance_bulk():
    """Mark attendance in bulk for multiple students.
    ---
    tags:
      - Attendance
    consumes:
      - application/json
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            student_ids:
              type: array
              items:
                type: string
              description: List of registration numbers
            status:
              type: string
              enum: [Present, Absent]
              default: Present
          required:
            - student_ids
    responses:
      200:
        description: Number of updated records
        schema:
          type: object
          properties:
            updated:
              type: integer
      400:
        description: Invalid request body
    """
    if not _check_rate_limit("attendance_bulk"):
      return jsonify({"error": "Rate limit exceeded."}), 429

    data = request.get_json(silent=True)
    if not data or "student_ids" not in data:
        return jsonify({"error": "Missing 'student_ids' in request body."}), 400

    reg_nos = data["student_ids"]
    status = data.get("status", "Present")
    if status not in ("Present", "Absent"):
        return jsonify({"error": "status must be 'Present' or 'Absent'."}), 400

    entries = []
    not_found = []
    for reg_no in reg_nos:
        student = database.get_student_by_reg_no(reg_no)
        if student is None:
            not_found.append(reg_no)
            continue
        entries.append({
            "student_id": student["_id"],
            "status": status,
            "confidence_score": 0.0,
        })

    updated = 0
    if entries:
        updated = database.bulk_upsert_attendance(entries)

    result = {"updated": updated}
    if not_found:
        result["not_found"] = not_found
    return jsonify(result)


# ── Student Self-Service Portal ───────────────────────────────────────────

@bp.route("/student", methods=["GET", "POST"])
def student_portal():
    if request.method == "POST":
        # Login: validate registration number
        reg_no = sanitize_string(request.form.get("registration_number", "").strip())
        if not reg_no:
            flash("Please enter your registration number.", "danger")
            return render_template("student_login.html"), 400

        student = database.get_student_by_reg_no(reg_no)
        if student is None:
            flash("Registration number not found.", "danger")
            return render_template("student_login.html"), 404

        session["student_reg_no"] = reg_no
        return redirect(url_for("main.student_portal"))

    # GET: check if already logged in
    reg_no = session.get("student_reg_no")
    if not reg_no:
        return render_template("student_login.html")

    summary = database.get_student_attendance_summary(reg_no)
    if summary is None:
        # Student was deleted after login
        session.pop("student_reg_no", None)
        flash("Student record not found. Please log in again.", "danger")
        return render_template("student_login.html")

    return render_template("student_portal.html", student=summary)


@bp.route("/student/logout")
def student_logout():
    session.pop("student_reg_no", None)
    return redirect(url_for("main.student_portal"))


# ── At-Risk Students API ─────────────────────────────────────────────────

@bp.route("/api/at_risk")
def api_at_risk():
    """Return students whose attendance is below the configured threshold.
    ---
    tags:
      - Students
    parameters:
      - name: days
        in: query
        type: integer
        required: false
        default: 30
        description: Number of past days to evaluate
      - name: threshold
        in: query
        type: integer
        required: false
        description: Attendance percentage threshold (default from config)
    responses:
      200:
        description: List of at-risk student objects
        schema:
          type: array
          items:
            type: object
            properties:
              name:
                type: string
              reg_no:
                type: string
              percentage:
                type: number
              days_present:
                type: integer
              days_total:
                type: integer
    """
    days = request.args.get("days", 30, type=int)
    threshold = request.args.get("threshold", None, type=int)
    data = database.get_at_risk_students(days=days, threshold=threshold)
    return jsonify(data)


# ── REST API: Students ────────────────────────────────────────────────────

@bp.route("/api/students", methods=["GET"])
def api_students_list():
    """Return all registered students.
    ---
    tags:
      - Students
    responses:
      200:
        description: List of student objects
        schema:
          type: array
          items:
            type: object
            properties:
              _id:
                type: string
              name:
                type: string
              semester:
                type: integer
              registration_number:
                type: string
              section:
                type: string
    """
    students = database.get_all_students()
    for s in students:
        s["_id"] = str(s["_id"])
    return jsonify(students)


@bp.route("/api/students", methods=["POST"])
def api_students_create():
    """Create a new student record.
    ---
    tags:
      - Students
    consumes:
      - application/json
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            name:
              type: string
            semester:
              type: integer
            registration_number:
              type: string
            section:
              type: string
            image_paths:
              type: array
              items:
                type: string
              description: List of filesystem paths to face images
          required:
            - name
            - semester
            - registration_number
            - section
            - image_paths
    responses:
      201:
        description: Student created successfully
        schema:
          type: object
          properties:
            id:
              type: string
            name:
              type: string
            encodings_count:
              type: integer
      400:
        description: Validation error
      409:
        description: Duplicate registration number
    """
    if not _check_rate_limit("students_create"):
      return jsonify({"error": "Rate limit exceeded."}), 429

    data = request.get_json(silent=True)
    if not data:
      return jsonify({"error": "JSON body required."}), 400

    name = sanitize_string(data.get("name", ""))
    semester_raw = data.get("semester")
    reg_no = sanitize_string(data.get("registration_number", ""))
    section = sanitize_string(data.get("section", ""))
    image_paths = data.get("image_paths", [])

    errors: list[str] = []
    if not name:
      errors.append("name is required.")
    if not reg_no:
      errors.append("registration_number is required.")
    if not section:
      errors.append("section is required.")
    try:
      semester = int(semester_raw)
      if semester < 1 or semester > 12:
        errors.append("semester must be between 1 and 12.")
    except (ValueError, TypeError):
      errors.append("semester must be a valid integer.")
      semester = None
    if not image_paths:
      errors.append("image_paths is required (list of file paths).")

    if errors:
      return jsonify({"errors": errors}), 400

    encodings: list[np.ndarray] = []
    for i, path in enumerate(image_paths, 1):
      resolved = _resolve_upload_reference(str(path))
      if resolved is None or not os.path.isfile(resolved):
        errors.append(f"Image {i}: invalid file path.")
        continue

      try:
        img_bgr = cv2.imread(resolved)
        if img_bgr is None:
          errors.append(f"Image {i}: could not read image file.")
          continue

        ok, reason = check_image_quality(img_bgr)
        if not ok:
          errors.append(f"Image {i}: {reason}")
          continue

        enc = generate_encoding(resolved)
        encodings.append(enc)
      except ValueError as exc:
        errors.append(f"Image {i}: {exc}")
        continue

    if errors:
      return jsonify({"errors": errors}), 400

    if not encodings:
      return jsonify({"error": "No valid face encodings could be generated."}), 400

    try:
      student_id = database.insert_student(
        name,
        semester,
        reg_no,
        section,
        encodings,
      )
    except ValueError as exc:
      return jsonify({"error": str(exc)}), 409

    encoding_cache.refresh()
    return jsonify({
      "id": str(student_id),
      "name": name,
      "encodings_count": len(encodings),
    }), 201


@bp.route("/api/students/<reg_no>", methods=["GET"])
def api_student_detail(reg_no):
    """Return a single student by registration number.
    ---
    tags:
      - Students
    parameters:
      - name: reg_no
        in: path
        type: string
        required: true
        description: Student registration number
    responses:
      200:
        description: Student object
        schema:
          type: object
          properties:
            _id:
              type: string
            name:
              type: string
            semester:
              type: integer
            registration_number:
              type: string
            section:
              type: string
      404:
        description: Student not found
    """
    student = database.get_student_by_reg_no(reg_no)
    if student is None:
        return jsonify({"error": "Student not found."}), 404

    student["_id"] = str(student["_id"])
    # Remove binary encoding data from response
    student.pop("encodings", None)
    student.pop("face_encoding", None)
    student.pop("created_at", None)
    return jsonify(student)


@bp.route("/api/students/<reg_no>", methods=["DELETE"])
def api_student_delete(reg_no):
    """Delete a student and all their attendance records.
    ---
    tags:
      - Students
    parameters:
      - name: reg_no
        in: path
        type: string
        required: true
        description: Student registration number
    responses:
      200:
        description: Student deleted successfully
        schema:
          type: object
          properties:
            deleted:
              type: boolean
            registration_number:
              type: string
      404:
        description: Student not found
    """
    deleted = database.delete_student(reg_no)
    if not deleted:
        return jsonify({"error": "Student not found."}), 404

    encoding_cache.refresh()
    return jsonify({"deleted": True, "registration_number": reg_no})


# ── REST API: Attendance ──────────────────────────────────────────────────

@bp.route("/api/attendance", methods=["GET"])
def api_attendance_list():
    """Return attendance records with optional filters.
    ---
    tags:
      - Attendance
    parameters:
      - name: date
        in: query
        type: string
        required: false
        description: Filter by date (YYYY-MM-DD)
      - name: reg_no
        in: query
        type: string
        required: false
        description: Filter by student registration number
      - name: start_date
        in: query
        type: string
        required: false
        description: Start of date range (YYYY-MM-DD)
      - name: end_date
        in: query
        type: string
        required: false
        description: End of date range (YYYY-MM-DD)
    responses:
      200:
        description: List of attendance records
        schema:
          type: array
          items:
            type: object
            properties:
              name:
                type: string
              registration_number:
                type: string
              section:
                type: string
              semester:
                type: integer
              date:
                type: string
              time:
                type: string
              status:
                type: string
              confidence_score:
                type: number
      400:
        description: Invalid parameters
    """
    reg_no = request.args.get("reg_no", "").strip()
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    date = request.args.get("date", "").strip()

    if reg_no:
        records = database.get_attendance_by_student(reg_no)
    elif start_date and end_date:
        sd, err1 = _validate_date_param(start_date, "start_date")
        ed, err2 = _validate_date_param(end_date, "end_date")
        if err1 or err2:
            return jsonify({"error": err1 or err2}), 400
        if sd > ed:
            return jsonify({"error": "start_date must be <= end_date."}), 400
        records = database.get_attendance_by_date_range(start_date, end_date)
    elif date:
        _, err = _validate_date_param(date, "date")
        if err:
            return jsonify({"error": err}), 400
        records = database.get_attendance(date)
    else:
        records = database.get_attendance()

    return jsonify(records)


@bp.route("/api/attendance", methods=["POST"])
def api_attendance_mark():
    """Manually mark attendance for a student.
    ---
    tags:
      - Attendance
    consumes:
      - application/json
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            reg_no:
              type: string
              description: Student registration number
            status:
              type: string
              enum: [Present, Absent]
              default: Present
          required:
            - reg_no
    responses:
      200:
        description: Attendance marked result
        schema:
          type: object
          properties:
            marked:
              type: boolean
            reg_no:
              type: string
            status:
              type: string
      400:
        description: Invalid request body
      404:
        description: Student not found
    """
    if not _check_rate_limit("attendance_mark"):
      return jsonify({"error": "Rate limit exceeded."}), 429

    data = request.get_json(silent=True)
    if not data or "reg_no" not in data:
        return jsonify({"error": "Missing 'reg_no' in request body."}), 400

    reg_no = sanitize_string(data["reg_no"])
    status = data.get("status", "Present")

    if status not in ("Present", "Absent"):
        return jsonify({"error": "status must be 'Present' or 'Absent'."}), 400

    student = database.get_student_by_reg_no(reg_no)
    if student is None:
        return jsonify({"error": "Student not found."}), 404

    entries = [{
        "student_id": student["_id"],
        "status": status,
        "confidence_score": 0.0,
    }]
    count = database.bulk_upsert_attendance(entries)

    return jsonify({
        "marked": count > 0,
        "reg_no": reg_no,
        "status": status,
    })


# ── Celery Task Status ────────────────────────────────────────────────────

@bp.route("/api/task/<task_id>")
def api_task_status(task_id):
    """Check the status of an asynchronous Celery task.
    ---
    tags:
      - Tasks
    parameters:
      - name: task_id
        in: path
        type: string
        required: true
        description: Celery task ID
    responses:
      200:
        description: Task state and result
        schema:
          type: object
          properties:
            state:
              type: string
            result:
              type: object
    """
    try:
        from celery_app import celery_app
    except ImportError:
        return jsonify({"error": "Celery is not configured."}), 503

    task = celery_app.AsyncResult(task_id)
    if task.state == "PENDING":
        return jsonify({"state": "PENDING"})
    elif task.state == "FAILURE":
        return jsonify({"state": "FAILURE", "error": str(task.result)})
    else:
        return jsonify({"state": task.state, "result": task.result})


# ── Async CSV Generation ─────────────────────────────────────────────────

@bp.route("/api/report/csv/async", methods=["POST"])
def api_report_csv_async():
    """Launch an asynchronous CSV generation task.
    ---
    tags:
      - Reports
    consumes:
      - application/json
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            date:
              type: string
              description: Single date filter (YYYY-MM-DD)
            reg_no:
              type: string
              description: Filter by student registration number
            start_date:
              type: string
              description: Start of date range (YYYY-MM-DD)
            end_date:
              type: string
              description: End of date range (YYYY-MM-DD)
            full:
              type: string
              description: Set to '1' to export entire history
    responses:
      202:
        description: Task launched
        schema:
          type: object
          properties:
            task_id:
              type: string
      503:
        description: Celery is not configured
    """
    try:
        from celery_app import generate_csv_task
    except ImportError:
        return jsonify({"error": "Celery is not configured."}), 503

    data = request.get_json(silent=True) or {}
    date = sanitize_string(str(data.get("date", "")).strip())
    reg_no = sanitize_string(str(data.get("reg_no", "")).strip())
    start_date = sanitize_string(str(data.get("start_date", "")).strip())
    end_date = sanitize_string(str(data.get("end_date", "")).strip())
    full = str(data.get("full", "")).strip()

    if full == "1":
      task = generate_csv_task.delay("full")
    elif reg_no:
      task = generate_csv_task.delay("student", reg_no=reg_no)
    elif start_date and end_date:
      task = generate_csv_task.delay(
        "range",
        start_date=start_date,
        end_date=end_date,
      )
    else:
      date_str = date or today_str()
      task = generate_csv_task.delay("date", date_str=date_str)
    return jsonify({"task_id": task.id}), 202
