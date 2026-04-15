"""Report page and export route registrations."""

import io

from flask import jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app_web.decorators import require_roles


def register_report_routes(bp):
    @bp.route("/report")
    @require_roles("admin", "teacher")
    def report():
        from app_web import routes as routes_module

        date = request.args.get("date", routes_module.today_str())
        records = routes_module.database.get_attendance(date)
        return render_template(
            "report.html",
            records=records,
            selected_date=date,
        )

    @bp.route("/report/csv")
    @require_roles("admin", "teacher")
    def report_csv():
        from app_web import routes as routes_module

        filters, error = routes_module._parse_report_filters(
            dict(request.args),
            default_date=routes_module.today_str(),
        )
        if error is not None:
            payload, status_code = error
            return routes_module._api_error(payload["error"], status_code)

        mode = filters["mode"]
        if mode == "full":
            df = routes_module.database.get_attendance_csv_full()
            filename = "attendance_full_history.csv"
        elif mode == "student":
            reg_no = filters["reg_no"]
            df = routes_module.database.get_attendance_csv_by_student(reg_no)
            filename = f"attendance_{reg_no}.csv"
        elif mode == "range":
            start_date = filters["start_date"]
            end_date = filters["end_date"]
            df = routes_module.database.get_attendance_csv_by_date_range(start_date, end_date)
            filename = f"attendance_{start_date}_to_{end_date}.csv"
        else:
            date = filters["date"]
            df = routes_module.database.get_attendance_csv(date)
            filename = f"attendance_{date}.csv"

        buf = io.BytesIO()
        df.to_csv(buf, index=False)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename,
        )

    @bp.route("/report/xlsx")
    @require_roles("admin", "teacher")
    def report_xlsx():
        from app_web import routes as routes_module

        filters, error = routes_module._parse_report_filters(
            dict(request.args),
            default_date=routes_module.today_str(),
        )
        if error is not None:
            payload, status_code = error
            return routes_module._api_error(payload["error"], status_code)

        mode = filters["mode"]
        if mode == "full":
            df = routes_module.database.get_attendance_csv_full()
            filename = "attendance_full_history.xlsx"
        elif mode == "student":
            reg_no = filters["reg_no"]
            df = routes_module.database.get_attendance_csv_by_student(reg_no)
            filename = f"attendance_{reg_no}.xlsx"
        elif mode == "range":
            start_date = filters["start_date"]
            end_date = filters["end_date"]
            df = routes_module.database.get_attendance_csv_by_date_range(start_date, end_date)
            filename = f"attendance_{start_date}_to_{end_date}.xlsx"
        else:
            date = filters["date"]
            df = routes_module.database.get_attendance_csv(date)
            filename = f"attendance_{date}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        columns = list(df.columns)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        ws.freeze_panes = "A2"

        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    @bp.route("/api/report/csv/async", methods=["POST"])
    @require_roles("admin", "teacher")
    def api_report_csv_async():
        from app_web import routes as routes_module

        if not routes_module._check_rate_limit("report_csv_async"):
            return routes_module._api_error("Rate limit exceeded.", 429)

        try:
            from celery_app import generate_csv_task
        except ImportError:
            return routes_module._api_error("Celery is not configured.", 503)

        data = request.get_json(silent=True) or {}
        filters, error = routes_module._parse_report_filters(
            data,
            default_date=routes_module.today_str(),
        )
        if error is not None:
            payload, status_code = error
            return routes_module._api_error(payload["error"], status_code)

        mode = filters["mode"]
        if mode == "full":
            task = generate_csv_task.delay("full")
        elif mode == "student":
            reg_no = filters["reg_no"]
            task = generate_csv_task.delay("student", reg_no=reg_no)
        elif mode == "range":
            start_date = filters["start_date"]
            end_date = filters["end_date"]
            task = generate_csv_task.delay(
                "range",
                start_date=start_date,
                end_date=end_date,
            )
        else:
            date_str = filters["date"]
            task = generate_csv_task.delay("date", date_str=date_str)

        return jsonify({"task_id": task.id}), 202